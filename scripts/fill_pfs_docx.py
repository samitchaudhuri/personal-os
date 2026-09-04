#!/usr/bin/env python3
"""Fill Amerifund PFS Word doc from Chaudhuri_PFS_Source.xlsx (surgical XML).

Reads PFS Map, Schedules, and Meta tabs; inserts values into Chaudhuri_PFS_Template.docx
(empty layout master) by w14:paraId without reserializing document.xml (preserves Word namespaces).

Requires: openpyxl — see scripts/requirements.txt; run via scripts/.venv
  (python3 -m venv scripts/.venv && scripts/.venv/bin/pip install -r scripts/requirements.txt)

Example:
  python scripts/fill_pfs_docx.py \\
    --xlsx "$PFS_DIR/Chaudhuri_PFS_Source.xlsx" \\
    --template "$PFS_DIR/Chaudhuri_PFS_Template.docx" \\
    --output "$PFS_DIR/Chaudhuri_PFS_2026-06-24.docx" \\
    --date 6/23/2026

The script reads the empty layout master Chaudhuri_PFS_Template.docx (never write
script output back to that file). See vault/Agent/Workflows/Personal Financial Statement (PFS).md.
"""

from __future__ import annotations

import argparse
import re
import sys
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.sax.saxutils import escape

try:
    import openpyxl
except ImportError:
    print("fill_pfs_docx.py requires openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Reached via personal-os/gdrive (repo symlink) rather than the absolute
# CloudStorage path, so this works across machines/clones without editing.
# Override with --pfs-dir if the gdrive symlink is not present on this machine.
DEFAULT_PFS_DIR = (
    Path(__file__).resolve().parent.parent
    / "gdrive" / "private" / "ULC-personal" / "Personal" / "Personal Financial Statements"
)


def fmt_money(n) -> str:
    if n is None or n == "":
        return ""
    return f"{int(round(float(n))):,}"


def fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return f"{val.strftime('%b')} {val.year}"
    if isinstance(val, date):
        return f"{val.strftime('%b')} {val.year}"
    return str(val)


def load_map(wb) -> dict:
    return {
        row[0]: row[2]
        for row in wb["PFS Map"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }


def load_schedules(wb) -> dict:
    sched: dict = {}
    for row in wb["Schedules"].iter_rows(min_row=2, values_only=True):
        sch, r, field, val = row
        if sch is None:
            continue
        sched.setdefault(str(sch).strip(), {}).setdefault(int(r), {})[field] = val
    return sched


def load_meta(wb) -> dict:
    return {
        row[0]: row[1]
        for row in wb["Meta"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }


def run_xml(text: str) -> str:
    return (
        f'<w:r><w:rPr><w:sz w:val="18"/></w:rPr>'
        f"<w:t>{escape(str(text))}</w:t></w:r>"
    )


def run_xml_bold(text: str) -> str:
    """Header line run — matches template (bold, 9pt / w:sz 18 half-points)."""
    space_attr = ""
    if text and (text[0] == " " or text[-1] == " "):
        space_attr = ' xml:space="preserve"'
    return (
        f'<w:r><w:rPr><w:b/><w:sz w:val="18"/></w:rPr>'
        f"<w:t{space_attr}>{escape(str(text))}</w:t></w:r>"
    )


def run_xml_bold_tab_then_text(after_tab: str) -> str:
    """Single run with Word tab stop then text (template date-line pattern)."""
    return (
        f'<w:r><w:rPr><w:b/><w:sz w:val="18"/></w:rPr>'
        f"<w:tab/><w:t>{escape(after_tab)}</w:t></w:r>"
    )


def insert_run_by_paraid(xml: str, para_id: str, text) -> str:
    if text is None or text == "":
        return xml
    run = run_xml(text)
    pattern = rf'(<w:p[^>]*w14:paraId="{para_id}"[^>]*>.*?)(</w:p>)'
    new_xml, count = re.subn(
        pattern,
        lambda m: m.group(1) + run + m.group(2),
        xml,
        count=1,
        flags=re.DOTALL,
    )
    if count != 1:
        raise ValueError(f"paraId not found in template: {para_id} ({text!r})")
    return new_xml


def fix_date_paragraph(xml: str, date_str: str) -> str:
    pattern = rf'(<w:p[^>]*w14:paraId="0274BF59"[^>]*>)(.*?)(</w:p>)'

    def repl(m: re.Match) -> str:
        inner = m.group(2)
        ppr = re.search(r"(<w:pPr>.*?</w:pPr>)", inner, re.DOTALL)
        ppr_xml = ppr.group(1) if ppr else ""
        # Match template: title | [tab] Date: | date value (bold; w:tab not literal \t)
        runs = (
            run_xml_bold("Personal Financial Statement")
            + run_xml_bold_tab_then_text("Date:")
            + run_xml_bold(f" {date_str}")
        )
        return m.group(1) + ppr_xml + runs + m.group(3)

    new_xml, count = re.subn(pattern, repl, xml, count=1, flags=re.DOTALL)
    if count != 1:
        raise ValueError("Date paragraph (paraId 0274BF59) not found in template")
    return new_xml


def build_fills(pfs: dict, sched: dict, meta: dict) -> dict[str, str]:
    c, d = sched["C"][1], sched["D"][1]
    salary = fmt_money(meta.get("Salary_Bonus_Commissions", 280000))

    fills = {
        "79F59BC8": fmt_money(pfs["Cash on Hand"]),
        "15750DF2": fmt_money(pfs["Notes Payable to Banks"]),
        "4DC07A2E": fmt_money(pfs["Marketable Securities — Schedule A"]),
        "134954DE": fmt_money(pfs["Partnership Interests — Schedule C"]),
        "20D2F6BD": fmt_money(pfs["Real Estate Owned — Schedule D"]),
        "170CBFD5": fmt_money(pfs["Real Estate Mortgages Payable"]),
        "44D57A1F": fmt_money(pfs["Retirement Account(s)"]),
        "36A11C5C": f"{fmt_money(pfs['Other Debts — Itemize'])} (credit cards)",
        "1CECF511": fmt_money(pfs["Automobiles and Other Personal Property"]),
        "6AB19687": fmt_money(pfs["Misc. Household goods"]),
        "1178FDA3": fmt_money(pfs["Total Liabilities"]),
        "488F0EF8": fmt_money(pfs["Net Worth"]),
        "729CEA70": fmt_money(pfs["Total Assets"]),
        "4AC3744D": fmt_money(pfs["Total Assets"]),
        "7507772B": fmt_money(sched["A"][1]["Market_Value"]),
        "62C28C5B": c["Name"],
        "5C8F5826": c["Owner"],
        "02358286": str(int(c["Pct_Ownership"])),
        "59DD048D": c["Type"],
        "0F048612": str(int(c["Year"])),
        "3B9819C9": fmt_money(c["Cost"]),
        "5522052A": fmt_money(c["Market_Value"]),
        "152D1199": fmt_money(c["Mortgage_or_Loan"]),
        "7D08DF4A": fmt_money(c["Equity"]),
        "2BFA2BD1": fmt_money(d["Cost"]),
        "376FB616": fmt_money(d["Market_Value"]),
        "53A93691": fmt_money(d["Mortgage"]),
        "1D9F4AF2": str(int(d["Annual_Net_Cash_Flow"])),
        "030AD978": fmt_money(d["Annual_Payment"]),
        "0B66292C": salary,
        "02682C9D": salary,
        "11305FB7": salary,
        "77CECE01": "0",
        "0D0519A2": "0",
        "18D8D488": sched["F"][1]["Purpose"],
        "515E1429": fmt_date(sched["F"][1]["Original_Date"]),
        "21384FEC": fmt_money(sched["F"][1]["High_Credit"]),
        "085EE136": fmt_money(sched["F"][1]["Owe_Currently"]),
        "3AF28B3E": sched["F"][1]["Secured"],
        "4D89CC96": sched["F"][2]["Purpose"],
        "3A7A67CA": fmt_date(sched["F"][2]["Original_Date"]),
        "04EC8D4C": fmt_money(sched["F"][2]["High_Credit"]),
        "01F55915": fmt_money(sched["F"][2]["Owe_Currently"]),
        "3C671E25": sched["F"][2]["Secured"],
        "6A01DE43": fmt_money(pfs["Notes Payable to Banks"]),
    }

    e_rows = [
        ("22A9A9AC", "61DEE780", "3F64064C", "7B52618E", "22757D86", "105D379A", 1),
        ("7CCA3FE0", "760162D1", "2E38E217", "57465110", "3FA0C24E", "154B0EF8", 2),
        ("75FE9D75", "43EB95C1", "3FDCD784", "1644CAF5", "682AD0F9", "4D9FD498", 3),
        ("62071E2B", "461D3641", "69B74543", "36E8F987", "22B96FE2", "3F0601A6", 4),
    ]
    for pids in e_rows:
        e = sched["E"][pids[-1]]
        vals = [
            fmt_money(e["Face_Amount"]),
            e["Company"],
            e["Owner"],
            e["Beneficiary"],
            fmt_money(e["CSV"]),
            fmt_money(e["Loans"]),
        ]
        for pid, val in zip(pids[:-1], vals):
            fills[pid] = val

    return fills


def fill_docx(
    xlsx_path: Path,
    template_path: Path,
    output_path: Path,
    date_str: str,
) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    pfs = load_map(wb)
    sched = load_schedules(wb)
    meta = load_meta(wb)
    fills = build_fills(pfs, sched, meta)

    with zipfile.ZipFile(template_path, "r") as zin:
        items = [(i.filename, i, zin.read(i.filename)) for i in zin.infolist()]
        xml = next(data.decode("utf-8") for fn, _, data in items if fn == "word/document.xml")

    xml = fix_date_paragraph(xml, date_str)
    for para_id, value in fills.items():
        xml = insert_run_by_paraid(xml, para_id, value)

    if "xmlns:wpc=" not in xml[:3000]:
        raise RuntimeError("document.xml namespaces lost — wrong template or corrupt edit")

    tmp = output_path.with_suffix(".tmp.docx")
    with zipfile.ZipFile(tmp, "w") as zout:
        for fn, info, data in items:
            payload = xml.encode("utf-8") if fn == "word/document.xml" else data
            zout.writestr(info, payload)
    tmp.replace(output_path)

    print(f"Wrote {output_path}")
    print(f"  Total Assets:     {fmt_money(pfs['Total Assets'])}")
    print(f"  Total Liabilities:{fmt_money(pfs['Total Liabilities'])}")
    print(f"  Net Worth:        {fmt_money(pfs['Net Worth'])}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", type=Path, help="Chaudhuri_PFS_Source.xlsx export")
    p.add_argument(
        "--template",
        type=Path,
        help="Empty layout master (Chaudhuri_PFS_Template.docx — script never writes here)",
    )
    p.add_argument("--output", type=Path, help="Output docx path")
    p.add_argument(
        "--date",
        help="PFS header date, e.g. 6/23/2026 (default: Meta PFS_Date_Header or today)",
    )
    p.add_argument(
        "--pfs-dir",
        type=Path,
        default=DEFAULT_PFS_DIR,
        help=f"PFS Drive folder (default: {DEFAULT_PFS_DIR})",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    pfs_dir = args.pfs_dir
    xlsx = args.xlsx or pfs_dir / "Chaudhuri_PFS_Source.xlsx"
    template = args.template or pfs_dir / "Chaudhuri_PFS_Template.docx"
    output = args.output or pfs_dir / "Chaudhuri_PFS_2026-06-24.docx"

    if args.date:
        date_str = args.date
    else:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        meta = load_meta(wb)
        hdr = meta.get("PFS_Date_Header")
        if isinstance(hdr, datetime):
            date_str = f"{hdr.month}/{hdr.day}/{hdr.year}"
        elif isinstance(hdr, date):
            date_str = f"{hdr.month}/{hdr.day}/{hdr.year}"
        else:
            today = date.today()
            date_str = f"{today.month}/{today.day}/{today.year}"

    for path, label in [(xlsx, "xlsx"), (template, "template docx")]:
        if not path.exists():
            print(f"Missing {label}: {path}", file=sys.stderr)
            sys.exit(1)

    fill_docx(xlsx, template, output, date_str)


if __name__ == "__main__":
    main()
